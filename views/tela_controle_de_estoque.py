import sys
import os
from PySide6.QtWidgets import (
    QApplication, QMainWindow, QWidget, QLabel,
    QVBoxLayout, QPushButton, QTableWidget,
    QTableWidgetItem, QLineEdit, QFormLayout,
    QDialog, QMessageBox, QHBoxLayout
)
from PySide6.QtCore import Qt
from openpyxl import Workbook, load_workbook

ARQUIVO_EXCEL = "produtos.xlsx"


# ===================== GERENCIADOR DE PRODUTOS =====================
class ProdutoDB:
    def __init__(self, arquivo):
        self.arquivo = arquivo
        if not os.path.exists(self.arquivo):
            self.criar_arquivo()

    def criar_arquivo(self):
        wb = Workbook()
        ws = wb.active
        ws.append(["codigo", "nome", "quantidade", "custo", "venda"])
        wb.save(self.arquivo)

    def carregar(self):
        wb = load_workbook(self.arquivo)
        ws = wb.active
        produtos = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            produtos.append({
                "codigo": row[0],
                "nome": row[1],
                "quantidade": row[2],
                "custo": row[3],
                "venda": row[4]
            })
        return produtos

    def salvar_todos(self, produtos):
        wb = Workbook()
        ws = wb.active
        ws.append(["codigo", "nome", "quantidade", "custo", "venda"])
        for p in produtos:
            ws.append([
                p["codigo"], p["nome"], p["quantidade"],
                p["custo"], p["venda"]
            ])
        wb.save(self.arquivo)

    def buscar_por_codigo(self, codigo):
        for p in self.carregar():
            if p["codigo"] == codigo:
                return p
        return None

    def buscar_por_nome(self, nome):
        resultados = []
        for p in self.carregar():
            if nome.lower() in p["nome"].lower():
                resultados.append(p)
        return resultados

    def adicionar_ou_atualizar(self, produto, entrada=False):
        produtos = self.carregar()
        for p in produtos:
            if p["codigo"] == produto["codigo"]:
                if entrada:
                    p["quantidade"] += produto["quantidade"]
                else:
                    p.update(produto)
                self.salvar_todos(produtos)
                return
        produtos.append(produto)
        self.salvar_todos(produtos)


# ===================== TELA DE CADASTRO =====================
class CadastroProduto(QDialog):
    def __init__(self, db):
        super().__init__()
        self.db = db
        self.setWindowTitle("Cadastro / Alteração de Produto")
        self.setFixedSize(400, 300)

        layout = QFormLayout(self)

        self.codigo = QLineEdit()
        self.nome = QLineEdit()
        self.quantidade = QLineEdit()
        self.custo = QLineEdit()
        self.venda = QLineEdit()

        layout.addRow("Código:", self.codigo)
        layout.addRow("Nome:", self.nome)
        layout.addRow("Quantidade:", self.quantidade)
        layout.addRow("Valor de Custo:", self.custo)
        layout.addRow("Valor de Venda:", self.venda)

        btn_salvar = QPushButton("Salvar / Atualizar")
        btn_salvar.clicked.connect(self.salvar)
        layout.addRow(btn_salvar)

    def salvar(self):
        try:
            produto = {
                "codigo": self.codigo.text(),
                "nome": self.nome.text(),
                "quantidade": int(self.quantidade.text()),
                "custo": float(self.custo.text()),
                "venda": float(self.venda.text())
            }
        except ValueError:
            QMessageBox.warning(self, "Erro", "Preencha corretamente todos os campos numéricos")
            return

        self.db.adicionar_ou_atualizar(produto)
        QMessageBox.information(self, "OK", "Produto salvo com sucesso")
        self.close()


# ===================== TELA DE ENTRADA =====================
class EntradaProduto(QDialog):
    def __init__(self, db):
        super().__init__()
        self.db = db
        self.setWindowTitle("Entrada de Produto")
        self.setFixedSize(300, 200)

        layout = QFormLayout(self)

        self.codigo = QLineEdit()
        self.quantidade = QLineEdit()

        layout.addRow("Código:", self.codigo)
        layout.addRow("Quantidade Entrada:", self.quantidade)

        btn = QPushButton("Dar Entrada")
        btn.clicked.connect(self.entrada)
        layout.addRow(btn)

    def entrada(self):
        codigo = self.codigo.text()
        try:
            qtd = int(self.quantidade.text())
        except ValueError:
            QMessageBox.warning(self, "Erro", "Quantidade deve ser um número")
            return

        produto = self.db.buscar_por_codigo(codigo)
        if not produto:
            QMessageBox.warning(self, "Erro", "Produto não encontrado")
            return

        self.db.adicionar_ou_atualizar({"codigo": codigo, "quantidade": qtd}, entrada=True)
        QMessageBox.information(self, "OK", "Entrada realizada")
        self.close()


# ===================== TELA DE BUSCA =====================
class BuscaProduto(QDialog):
    def __init__(self, db):
        super().__init__()
        self.db = db
        self.setWindowTitle("Busca de Produto")
        self.setFixedSize(500, 400)

        layout = QVBoxLayout(self)

        hbox = QHBoxLayout()
        self.input_busca = QLineEdit()
        hbox.addWidget(self.input_busca)
        btn_buscar = QPushButton("Buscar")
        btn_buscar.clicked.connect(self.buscar)
        hbox.addWidget(btn_buscar)
        layout.addLayout(hbox)

        self.table = QTableWidget(0, 5)
        self.table.setHorizontalHeaderLabels(
            ["Código", "Nome", "Qtd", "Custo", "Venda"]
        )
        layout.addWidget(self.table)

    def buscar(self):
        termo = self.input_busca.text()
        resultados = []
        if termo.isdigit():
            p = self.db.buscar_por_codigo(termo)
            if p:
                resultados.append(p)
        resultados += self.db.buscar_por_nome(termo)

        self.table.setRowCount(0)
        for p in resultados:
            row = self.table.rowCount()
            self.table.insertRow(row)
            self.table.setItem(row, 0, QTableWidgetItem(p["codigo"]))
            self.table.setItem(row, 1, QTableWidgetItem(p["nome"]))
            self.table.setItem(row, 2, QTableWidgetItem(str(p["quantidade"])))
            self.table.setItem(row, 3, QTableWidgetItem(str(p["custo"])))
            self.table.setItem(row, 4, QTableWidgetItem(str(p["venda"])))


# ===================== JANELA PRINCIPAL =====================
class EstoqueApp(QMainWindow):
    def __init__(self):
        super().__init__()
        self.setWindowTitle("Controle de Estoque")
        self.setFixedSize(600, 400)

        self.db = ProdutoDB(ARQUIVO_EXCEL)

        central = QWidget()
        self.setCentralWidget(central)

        # Layout principal com margens
        layout = QVBoxLayout(central)
        layout.setContentsMargins(30, 30, 30, 30)
        layout.setSpacing(20)

        # Título
        lbl = QLabel("Controle de Estoque")
        lbl.setAlignment(Qt.AlignCenter)
        lbl.setStyleSheet("font-size:22px; font-weight:bold;")
        layout.addWidget(lbl)

        # Layout de botões em grade (2 linhas x 2 colunas)
        grid = QHBoxLayout()
        layout.addLayout(grid)
        grid.setSpacing(15)

        # Botões com cores diferentes
        btn_cad = QPushButton("Cadastro / Alteração")
        btn_cad.setFixedHeight(50)
        btn_cad.setStyleSheet("background-color:#4CAF50; color:white; font-weight:bold;")
        btn_cad.clicked.connect(lambda: CadastroProduto(self.db).exec())

        btn_ent = QPushButton("Entrada de Produtos")
        btn_ent.setFixedHeight(50)
        btn_ent.setStyleSheet("background-color:#2196F3; color:white; font-weight:bold;")
        btn_ent.clicked.connect(lambda: EntradaProduto(self.db).exec())

        btn_bus = QPushButton("Buscar Produto")
        btn_bus.setFixedHeight(50)
        btn_bus.setStyleSheet("background-color:#FF9800; color:white; font-weight:bold;")
        btn_bus.clicked.connect(lambda: BuscaProduto(self.db).exec())

        btn_sair = QPushButton("Sair")
        btn_sair.setFixedHeight(50)
        btn_sair.setStyleSheet("background-color:#f44336; color:white; font-weight:bold;")
        btn_sair.clicked.connect(self.close)

        # Adicionar os botões à grade
        col1 = QVBoxLayout()
        col1.setSpacing(15)
        col1.addWidget(btn_cad)
        col1.addWidget(btn_ent)

        col2 = QVBoxLayout()
        col2.setSpacing(15)
        col2.addWidget(btn_bus)
        col2.addWidget(btn_sair)

        grid.addLayout(col1)
        grid.addLayout(col2)


# ===================== MAIN =====================
if __name__ == "__main__":
    app = QApplication(sys.argv)
    window = EstoqueApp()
    window.show()
    sys.exit(app.exec())
