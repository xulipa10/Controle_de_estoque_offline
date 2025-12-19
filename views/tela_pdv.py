import sys
import os
from datetime import datetime
from PySide6.QtWidgets import (
    QApplication, QMainWindow, QWidget, QLabel,
    QVBoxLayout, QHBoxLayout, QTableWidget, QTableWidgetItem,
    QLineEdit, QPushButton, QGridLayout, QMessageBox, QInputDialog
)
from PySide6.QtCore import Qt, QTimer
from PySide6.QtGui import QKeySequence, QShortcut
from openpyxl import load_workbook, Workbook

ARQUIVO_EXCEL = "Produtos.xlsx"

ARQUIVO_VENDAS = "Vendas.xlsx"


# ===================== GERENCIADOR DE PRODUTOS =====================
class ProdutoDB:
    def __init__(self, arquivo):
        self.arquivo = arquivo
        if not os.path.exists(self.arquivo):
            self.criar_arquivo()

    def criar_arquivo(self):
        wb = Workbook()
        ws = wb.active
        ws.append(["codigo", "nome", "quantidade", "custo", "venda"])
        wb.save(self.arquivo)

    def carregar(self):
        wb = load_workbook(self.arquivo)
        ws = wb.active
        produtos = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            produtos.append({
                "codigo": str(row[0]),
                "nome": row[1],
                "quantidade": row[2],
                "custo": row[3],
                "venda": row[4]
            })
        return produtos

    def buscar_por_codigo(self, codigo):
        for p in self.carregar():
            if p["codigo"] == codigo:
                return p
        return None




class VendaDB:
    def __init__(self, arquivo):
        self.arquivo = arquivo
        if not os.path.exists(self.arquivo):
            self.criar_arquivo()

    def criar_arquivo(self):
        wb = Workbook()

        ws_vendas = wb.active
        ws_vendas.title = "vendas"
        ws_vendas.append(["id", "data", "hora", "total", "pagamento"])

        ws_itens = wb.create_sheet("itens")
        ws_itens.append([
            "venda_id", "codigo", "descricao",
            "quantidade", "unitario", "total"
        ])

        wb.save(self.arquivo)

    def salvar_venda(self, total, pagamento, itens):
        wb = load_workbook(self.arquivo)
        ws_vendas = wb["vendas"]
        ws_itens = wb["itens"]

        venda_id = ws_vendas.max_row
        now = datetime.now()

        ws_vendas.append([
            venda_id,
            now.strftime("%d/%m/%Y"),
            now.strftime("%H:%M:%S"),
            total,
            pagamento
        ])

        for item in itens:
            ws_itens.append([
                venda_id,
                item["codigo"],
                item["descricao"],
                item["qtd"],
                item["unit"],
                item["total"]
            ])

        wb.save(self.arquivo)

# ===================== PDV =====================
class PDV(QMainWindow):
    def __init__(self):
        super().__init__()

        self.MODOS_PAGAMENTO = [
            "Dinheiro",
            "Cartão Débito",
            "Cartão Crédito",
            "PIX"
        ]

        self.pagamento_selecionado = None
        self.valor_recebido = 0.0
        self.troco = 0.0
        self.total_value = 0  # valor total da venda

        self.venda_db = VendaDB(ARQUIVO_VENDAS)
        self.db = ProdutoDB(ARQUIVO_EXCEL)

        self.setWindowTitle("PDV Expresso")
        self.resize(1366, 768)
        self.is_fullscreen = False

        central = QWidget()
        self.setCentralWidget(central)
        main_layout = QVBoxLayout(central)
        main_layout.setContentsMargins(2, 2, 2, 2)
        main_layout.setSpacing(4)

        # ================= Barra Superior =================
        self.top_bar = QLabel()
        self.top_bar.setFixedHeight(40)
        self.top_bar.setStyleSheet("""
            background-color: #e0e0e0;
            font-weight: bold;
            padding-left: 10px;
        """)
        main_layout.addWidget(self.top_bar)

        self.timer = QTimer(self)
        self.timer.timeout.connect(self.update_clock)
        self.timer.start(1000)
        self.update_clock()

        # ================= Área Central =================

        center_layout = QHBoxLayout()

        center_layout.addSpacing(150)

        # Tabela de Itens
        self.table = QTableWidget(0, 6)
        self.table.setHorizontalHeaderLabels(
            ["Item", "Código", "Descrição", "Qtd", "Unit.", "Total"]
        )
        self.table.verticalHeader().setVisible(False)
        self.table.setEditTriggers(QTableWidget.NoEditTriggers)
        self.table.setFocusPolicy(Qt.NoFocus)
        self.table.setStyleSheet("font-size: 14px;")
        # Depois de criar a tabela
        self.table.setColumnWidth(2, 350)  # pixels de largura para a coluna "Descrição"
        self.table.setColumnWidth(5, 110)

        center_layout.addWidget(self.table, 5)

        # Logo
        self.logo = QLabel("PDV\nEXPRESSO")
        self.logo.setAlignment(Qt.AlignCenter)
        self.logo.setStyleSheet("""
            font-size: 34px;
            font-weight: bold;
            color: #1e88e5;
        """)
        center_layout.addWidget(self.logo, 2)



        main_layout.addLayout(center_layout)

        # ================= Caixa de Código de Barras e Multiplicador =================
        barcode_layout = QHBoxLayout()
        barcode_layout.setContentsMargins(0, 0, 20, 0)  # margem direita
        barcode_layout.addStretch()  # empurra os widgets para a direita

        # Caixa de código de barras
        self.barcode_input = QLineEdit()
        self.barcode_input.setPlaceholderText("Digite ou passe o código de barras")
        self.barcode_input.setFixedHeight(40)
        self.barcode_input.setFixedWidth(200)
        self.barcode_input.setStyleSheet("font-size: 20px; padding-left: 10px;")
        self.barcode_input.returnPressed.connect(self.handle_barcode)
        barcode_layout.addWidget(self.barcode_input)

        # Conecta o evento de alteração de texto
        self.barcode_input.textChanged.connect(self.check_multiplier_inline)

        # Caixa de multiplicador
        self.qty_input = QLineEdit()
        self.qty_input.setFixedHeight(40)
        self.qty_input.setFixedWidth(60)
        self.qty_input.setText("1")  # valor padrão 1
        self.qty_input.setAlignment(Qt.AlignCenter)
        self.qty_input.setStyleSheet("font-size: 20px;")
        barcode_layout.addWidget(self.qty_input)

        main_layout.addLayout(barcode_layout)

        # ================= Total da Venda =================
        self.total_value = 0.0
        self.total_label = QLabel("TOTAL: R$ 0,00")
        self.total_label.setFixedHeight(80)  # maior que antes
        self.total_label.setAlignment(Qt.AlignRight | Qt.AlignVCenter)
        self.total_label.setStyleSheet("""
            font-size: 50px;
            font-weight: bold;
            color: #1e88e5;
            padding-right: 20px;
            background-color: #e0e0e0;
            border-radius: 8px;
        """)
        main_layout.addWidget(self.total_label)

        # ================= Barra de Funções =================
        func_layout = QGridLayout()
        func_layout.setSpacing(4)
        self.functions = [
            "F1 Ajuda", "F2 Cliente", "F3 Produto",
            "F4 Cancelar", "F5 Desconto", "F6 Pagamento",
            "F7", "F8", "F9",
            "F10", "F11 Tela Cheia", "F12 Finalizar"
        ]
        for col, text in enumerate(self.functions):
            btn = QPushButton(text)
            btn.setMinimumHeight(45)
            btn.setFocusPolicy(Qt.NoFocus)
            btn.setStyleSheet("""
                background-color: #cfe2ff;
                font-weight: bold;
            """)
            # conecta botão ao handle_function
            btn.clicked.connect(lambda checked, i=col+1: self.handle_function(i))
            func_layout.addWidget(btn, 0, col)
        main_layout.addLayout(func_layout)

        # ================= Atalhos F1-F12 =================
        for i in range(1, 13):
            shortcut = QShortcut(QKeySequence(f"F{i}"), self)
            shortcut.activated.connect(self.make_function_handler(i))

    def make_function_handler(self, func_number):
        return lambda: self.handle_function(func_number)

    def update_display_total(self):
        if self.pagamento_selecionado == "Dinheiro" and self.valor_recebido > 0:
            diferenca = self.valor_recebido - self.total_value

            if diferenca < 0:
                falta = abs(diferenca)
                valor = f"{falta:,.2f}".replace(",", "X").replace(".", ",").replace("X", ".")
                self.total_label.setText(f"FALTA: R$ {valor}")
                self.total_label.setStyleSheet("""
                    font-size: 50px;
                    font-weight: bold;
                    color: #d32f2f;  /* vermelho */
                    padding-right: 20px;
                    background-color: #e0e0e0;
                    border-radius: 8px;
                """)
            else:
                troco = diferenca
                valor = f"{troco:,.2f}".replace(",", "X").replace(".", ",").replace("X", ".")
                self.total_label.setText(f"TROCO: R$ {valor}")
                self.total_label.setStyleSheet("""
                    font-size: 50px;
                    font-weight: bold;
                    color: #2e7d32;  /* verde */
                    padding-right: 20px;
                    background-color: #e0e0e0;
                    border-radius: 8px;
                """)
        else:
            self.total_label.setStyleSheet("""
                font-size: 50px;
                font-weight: bold;
                color: #1e88e5;
                padding-right: 20px;
                background-color: #e0e0e0;
                border-radius: 8px;
            """)
            self.update_total()

    # ================= Verifica multiplicador inline =================
    def check_multiplier_inline(self, text):
        text = text.strip()
        if text.endswith('*'):
            try:
                qty = int(text[:-1])
                if qty > 0:
                    self.qty_input.setText(str(qty))
            except ValueError:
                self.qty_input.setText("1")
            self.barcode_input.clear()

    def handle_barcode(self):
        code = self.barcode_input.text().strip()

        if self.venda_bloqueada():
            QMessageBox.warning(
                self,
                "Pagamento em andamento",
                "Finalize ou cancele o pagamento para continuar"
            )
            self.barcode_input.clear()
            return

        if not code:
            return

        try:
            qty = int(self.qty_input.text())
            if qty < 1:
                qty = 1
        except ValueError:
            qty = 1

        produto = self.db.buscar_por_codigo(code)
        if produto:
            self.add_item(produto["codigo"], produto["nome"], qty, produto["venda"])
        else:
            QMessageBox.warning(self, "Produto não encontrado", f"Código: {code}")

        self.barcode_input.clear()
        self.qty_input.setText("1")  # reseta multiplicador

    # ================= Lógica de Adição/Cancelamento =================
    def add_item(self, code, desc, qty, price):

        row = self.table.rowCount()
        total = qty * price

        if self.venda_bloqueada():
            QMessageBox.warning(
                self,
                "Pagamento em andamento",
                "Não é possível adicionar itens após iniciar o pagamento"
            )
            return

        self.table.insertRow(row)
        self.table.setItem(row, 0, QTableWidgetItem(str(row + 1)))  # Número do item
        self.table.setItem(row, 1, QTableWidgetItem(code))
        self.table.setItem(row, 2, QTableWidgetItem(desc))
        self.table.setItem(row, 3, QTableWidgetItem(str(qty)))
        self.table.setItem(row, 4, QTableWidgetItem(f"{price:.2f}"))
        self.table.setItem(row, 5, QTableWidgetItem(f"{total:.2f}"))
        for i in range(0,6):
            if i != 2:
                self.table.item(row,i).setTextAlignment(Qt.AlignCenter)

        self.total_value += total
        self.update_display_total()


    def cancel_item_by_number(self, item_number):
        if self.venda_bloqueada():
            QMessageBox.warning(
                self,
                "Pagamento em andamento",
                "Não é possível cancelar itens após iniciar o pagamento"
            )
            return
        for row in range(self.table.rowCount()):
            if self.table.item(row, 0).text() == str(item_number):
                value = float(self.table.item(row, 5).text())
                self.total_value -= value
                self.table.removeRow(row)
                self.update_display_total()
                self.renumber_items()
                self.show_message(f"Item {item_number} cancelado")
                return
        self.show_message(f"Item {item_number} não encontrado")

    def renumber_items(self):
        for row in range(self.table.rowCount()):
            self.table.setItem(row, 0, QTableWidgetItem(str(row + 1)))

    def venda_bloqueada(self):
        if self.pagamento_selecionado != "Dinheiro":
            return False
        return self.valor_recebido >= self.total_value

    def update_total(self):

        self.total_label.setStyleSheet("""
                font-size: 50px;
                font-weight: bold;
                color: #1e88e5;
                padding-right: 20px;
                background-color: #e0e0e0;
                border-radius: 8px;
            """)

        text = f"{self.total_value:,.2f}"
        text = text.replace(",", "X").replace(".", ",").replace("X", ".")
        self.total_label.setText(f"TOTAL: R$ {text}")

    def update_clock(self):
        now = datetime.now().strftime("%d/%m/%Y %H:%M:%S")
        self.top_bar.setText(f"CAIXA LIVRE | OPERADOR: ADMIN | {now}")

    def show_message(self, message):
        self.top_bar.setText(f"{message} | {datetime.now().strftime('%H:%M:%S')}")

    def reset_venda(self):
        self.table.setRowCount(0)
        self.total_value = 0
        self.valor_recebido = 0
        self.pagamento_selecionado = None

        self.total_label.setStyleSheet("""
            font-size: 50px;
            font-weight: bold;
            color: #1e88e5;
            padding-right: 20px;
            background-color: #e0e0e0;
            border-radius: 8px;
        """)
        self.update_total()

    def toggle_fullscreen(self):
        if self.is_fullscreen:
            self.showNormal()
        else:
            self.showFullScreen()
        self.is_fullscreen = not self.is_fullscreen

    # ================= Funções F1-F12 =================
    def handle_function(self, func):

        if func == 1:
            self.show_message("Ajuda acionada")

        elif self.venda_bloqueada() and func not in (6, 12, 4):
            self.show_message("Pagamento em andamento")
            return

        elif func == 3:
            self.add_item("123", "Produto Exemplo", 1, 10.00)
        elif func == 4:
            if self.table.rowCount() == 0:
                self.show_message("Nenhum item para cancelar")
                return
            item_number, ok = QInputDialog.getInt(
                self, "Cancelar Item", "Digite o número do item a cancelar:", 1, 1, self.table.rowCount()
            )
            if ok:
                self.cancel_item_by_number(item_number)
        elif func == 6:
            if self.table.rowCount() == 0:
                self.show_message("Nenhuma venda em andamento")
                return

            pagamento, ok = QInputDialog.getItem(
                self,
                "Forma de Pagamento",
                "Selecione o pagamento:",
                self.MODOS_PAGAMENTO,
                0,
                False
            )

            if not ok:
                return

            self.pagamento_selecionado = pagamento
            self.troco = 0

            if pagamento == "Dinheiro":
                restante = self.total_value - self.valor_recebido

                valor, ok = QInputDialog.getDouble(
                    self,
                    "Pagamento em Dinheiro",
                    f"Total da venda: R$ {self.total_value:,.2f}\n"
                    f"Já recebido: R$ {self.valor_recebido:,.2f}\n"
                    f"Falta pagar: R$ {restante:,.2f}\n"
                    "Informe o valor recebido agora:",
                    0, 0, 999999, 2
                )

                if not ok:
                    return

                # ✅ soma apenas uma vez
                self.valor_recebido += valor

                # ✅ atualiza corretamente FALTA / TROCO
                self.update_display_total()

                self.show_message("Pagamento em dinheiro")




        elif func == 11:
            self.toggle_fullscreen()
        elif func == 12:

            if self.table.rowCount() == 0:
                self.show_message("Nenhuma venda para finalizar")
                return

            if not self.pagamento_selecionado:
                QMessageBox.warning(self, "Pagamento", "Selecione a forma de pagamento (F6)")
                return

            if self.pagamento_selecionado == "Dinheiro":
                if self.valor_recebido < self.total_value:
                    falta = self.total_value - self.valor_recebido
                    QMessageBox.warning(
                        self,
                        "Pagamento insuficiente",
                        f"Falta R$ {falta:,.2f} para concluir a venda"
                    )
                    return

            itens = []
            for row in range(self.table.rowCount()):
                itens.append({
                    "codigo": self.table.item(row, 1).text(),
                    "descricao": self.table.item(row, 2).text(),
                    "qtd": int(self.table.item(row, 3).text()),
                    "unit": float(self.table.item(row, 4).text()),
                    "total": float(self.table.item(row, 5).text())
                })

            self.venda_db.salvar_venda(
                total=self.total_value,
                pagamento=self.pagamento_selecionado,
                itens=itens
            )

            self.show_message("Venda finalizada com sucesso")

            self.reset_venda()
            self.show_message("Venda finalizada com sucesso")




        else:
            self.show_message(f"Função F{func}")

# ================= Inicialização =================
if __name__ == "__main__":
    app = QApplication(sys.argv)
    pdv = PDV()
    pdv.show()
    sys.exit(app.exec())
